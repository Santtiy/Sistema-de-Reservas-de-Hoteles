from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from reservations.models import Payment, Reservation
from rooms.models import Room


def get_kpis() -> dict:
    today = timezone.localdate()
    month_start = today.replace(day=1)
    month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)

    total_rooms = Room.objects.count()

    overlapping = Q(
        reservations__check_in__lt=today + timedelta(days=1),
        reservations__check_out__gt=today,
    ) & ~Q(reservations__status=Reservation.Status.CANCELLED)
    available_rooms_today = (
        Room.objects.filter(status=Room.Status.AVAILABLE)
        .exclude(overlapping)
        .distinct()
        .count()
    )

    active_statuses = [
        Reservation.Status.PENDING,
        Reservation.Status.CONFIRMED,
        Reservation.Status.CHECKED_IN,
    ]
    active_reservations = Reservation.objects.filter(status__in=active_statuses).count()

    reservations_this_month = Reservation.objects.filter(
        check_in__gte=month_start, check_in__lt=month_end
    ).count()

    revenue_this_month = (
        Payment.objects.filter(status=Payment.Status.COMPLETED, paid_at__isnull=False)
        .filter(paid_at__date__gte=month_start, paid_at__date__lt=month_end)
        .aggregate(total=Sum("amount"))
        .get("total")
        or 0
    )

    occupied_today = max(total_rooms - available_rooms_today, 0)
    occupancy_rate_today = (
        round((occupied_today / total_rooms) * 100, 2) if total_rooms else 0
    )

    User = get_user_model()
    new_users_this_month = User.objects.filter(
        date_joined__gte=month_start, date_joined__lt=month_end
    ).count()

    return {
        "total_rooms": total_rooms,
        "available_rooms_today": available_rooms_today,
        "active_reservations": active_reservations,
        "reservations_this_month": reservations_this_month,
        "revenue_this_month": revenue_this_month,
        "occupancy_rate_today": occupancy_rate_today,
        "new_users_this_month": new_users_this_month,
    }


def get_reservations_queryset(filters: dict):
    qs = Reservation.objects.select_related("guest", "room", "room__room_type")

    start: date | None = filters.get("fecha_inicio")
    end: date | None = filters.get("fecha_fin")
    status = filters.get("status")

    if start:
        qs = qs.filter(check_in__gte=start)
    if end:
        qs = qs.filter(check_out__lte=end)
    if status and status != "ALL":
        qs = qs.filter(status=status)

    return qs


def generate_pdf_report(
    title: str,
    headers: list[str],
    rows: list[list],
    filename: str,
    footer_summary: list[str] | None = None,
) -> HttpResponse:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    data = [headers] + rows
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#243042")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d6e2")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f8fb")]),
            ]
        )
    )
    story.append(table)

    if footer_summary:
        story.append(Spacer(1, 12))
        for line in footer_summary:
            story.append(Paragraph(line, styles["Normal"]))

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def generate_excel_report(
    title: str,
    headers: list[str],
    rows: list[list],
    filename: str,
) -> HttpResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reservas" if "reserva" in title.lower() else "Reporte"

    header_fill = PatternFill("solid", fgColor="243042")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.cell(row=1, column=1, value=title).font = Font(bold=True)

    header_row = 3
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_index, row in enumerate(rows, start=header_row + 1):
        for col_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.border = thin_border

    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column].width = max_length + 2

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response
